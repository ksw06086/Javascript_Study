import sys, os
import urllib3
from pprint import pprint
import datetime

sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
from openpyxl import load_workbook
import pandas as pd

# Excel 파일 불러오기
load_wb = load_workbook(r"C:\BackDev\Javascript_Study\상암지역 우주팀 12월 만남캘린더.xlsm")
load_ws = load_wb['데이터 입력 시트']

# 열 이름 가져오기 (첫 번째 행의 7번째 열부터)
columns = [cell.value for cell in next(load_ws.iter_rows(min_row=7, max_row=7, min_col=2))]

# 데이터 가져오기 (두 번째 행부터)
data = []
for row in load_ws.iter_rows(min_row=7, min_col=2):
    data.append([cell.value for cell in row])

f = pd.DataFrame(data, columns=columns)

#* 엑셀에서 데이터 가져옴
def get_datas():
    datas = []
    for i, data in enumerate(zip(f['만남날짜'], f['팀'], f['섭외유형'], f['오픈여부'], f['단계'], f['섭외자'], f['인도자'], f['만남자'], f['시간'], f['장소'], f['전티여부'], f['활용 전도의장'], f['결과'], f['결과 내용'], f['탈락사유'])):
        # print(f"{i} {url} {number}")
        date, team, types, openyn, level, sup, indo, manam, time, place, before_tel, use_place, res, res_body, fail_reason = data

        datas.append({
            'date' : date.strftime("%Y-%m-%d") if isinstance(date, datetime.datetime) else None, 
            'team' : team if team != None else '',
            'types' : types if types != None else '',
            'openyn' : openyn if openyn != None else '',
            'level' : level if level != None else '',
            'sup' : sup if sup != None else '',
            'indo' : indo if indo != None else '',
            'manam' : manam if manam != None else '',
            'time' : time if time != None else '',
            'place' : place if place != None else '',
            'before_tel' : before_tel if before_tel != None else '',
            'use_place' : use_place if use_place != None else '',
            'res' : res if res != None else '',
            'res_body' : res_body if res_body != None else '',
            'fail_reason' : fail_reason if fail_reason != None else ''
        })
    return [data for data in datas if(data['date'] != None)]

print(columns)
pprint(get_datas())